"""
Data file parsing service for XLSX, CSV, and JSON files.
Extracts headers, previews, and validates data for certificate generation.
"""

import csv
import json
import os


def parse_data_file(file_path):
    """
    Parse a data file and return structured results.

    Returns dict with:
        - headers: list of column names
        - rows: list of row dicts
        - row_count: total valid rows
        - skipped_rows: number of empty rows skipped
        - preview: first 3 rows as list of dicts
        - error: error message if applicable
    """
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".xlsx":
        return _parse_xlsx(file_path)
    elif ext == ".csv":
        return _parse_csv(file_path)
    elif ext == ".json":
        return _parse_json(file_path)
    else:
        return {"error": f"Unsupported file format: {ext}. Please upload an XLSX, CSV, or JSON file."}


def _parse_xlsx(file_path):
    """Parse an Excel file."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        rows_raw = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows_raw:
            return {"error": "Your data file appears to be empty. Please check and re-upload."}

        # First row = headers
        headers = [str(h).strip() if h is not None else f"Column_{i+1}" for i, h in enumerate(rows_raw[0])]

        # Check for completely empty headers
        if all(h.startswith("Column_") for h in headers):
            return {
                "error": "No headers detected in the first row. Please ensure your file has column headers.",
                "needs_header_confirmation": True,
                "raw_first_row": [str(v) for v in rows_raw[0] if v is not None],
            }

        rows = []
        skipped = 0

        for row_values in rows_raw[1:]:
            # Skip completely empty rows
            if all(v is None or str(v).strip() == "" for v in row_values):
                skipped += 1
                continue

            row_dict = {}
            for i, header in enumerate(headers):
                val = row_values[i] if i < len(row_values) else ""
                row_dict[header] = str(val) if val is not None else ""
            rows.append(row_dict)

        if not rows:
            return {"error": "Your data file appears to be empty. Please check and re-upload."}

        return {
            "headers": headers,
            "rows": rows,
            "row_count": len(rows),
            "skipped_rows": skipped,
            "preview": rows[:3],
        }

    except Exception as e:
        return {"error": f"Could not parse Excel file: {str(e)}"}


def _parse_csv(file_path):
    """Parse a CSV file."""
    try:
        # Detect encoding
        with open(file_path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames

            if not headers:
                return {"error": "Your data file appears to be empty. Please check and re-upload."}

            headers = [h.strip() for h in headers]
            rows = []
            skipped = 0

            for row in reader:
                # Skip empty rows
                if all(not v or v.strip() == "" for v in row.values()):
                    skipped += 1
                    continue

                clean_row = {}
                for h in headers:
                    val = row.get(h, "")
                    clean_row[h] = str(val).strip() if val else ""
                rows.append(clean_row)

        if not rows:
            return {"error": "Your data file appears to be empty. Please check and re-upload."}

        return {
            "headers": headers,
            "rows": rows,
            "row_count": len(rows),
            "skipped_rows": skipped,
            "preview": rows[:3],
        }

    except Exception as e:
        return {"error": f"Could not parse CSV file: {str(e)}"}


def _parse_json(file_path):
    """Parse a JSON file (expects array of objects)."""
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        if isinstance(data, dict):
            # Check if it has a "data" or "rows" key containing the array
            for key in ["data", "rows", "records", "items"]:
                if key in data and isinstance(data[key], list):
                    data = data[key]
                    break
            else:
                data = [data]  # Single object → list of one

        if not isinstance(data, list):
            return {"error": "JSON file must contain an array of objects or a single object."}

        if not data:
            return {"error": "Your data file appears to be empty. Please check and re-upload."}

        # Extract headers from all objects (union of keys)
        headers = []
        for item in data:
            if isinstance(item, dict):
                for key in item:
                    if key not in headers:
                        headers.append(key)

        rows = []
        skipped = 0

        for item in data:
            if not isinstance(item, dict):
                skipped += 1
                continue

            if all(not str(v).strip() for v in item.values()):
                skipped += 1
                continue

            row_dict = {}
            for h in headers:
                val = item.get(h, "")
                row_dict[h] = str(val) if val is not None else ""
            rows.append(row_dict)

        if not rows:
            return {"error": "Your data file appears to be empty. Please check and re-upload."}

        return {
            "headers": headers,
            "rows": rows,
            "row_count": len(rows),
            "skipped_rows": skipped,
            "preview": rows[:3],
        }

    except json.JSONDecodeError as e:
        return {"error": f"Invalid JSON file: {str(e)}"}
    except Exception as e:
        return {"error": f"Could not parse JSON file: {str(e)}"}
